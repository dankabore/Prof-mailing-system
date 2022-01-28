"""
For this project, you first need in order to use the openpyxl module to install it. How do you do that, first go to your idle terminator whether it is Pycharm or cmd or whatever that is,
go there and type the following commande: pip install openpyxl
This should download the module and install it into your idle. Then you are ready to start. Good luck!!
NB: You can download and edit the exel document below according to your needs. If you change the name of the file or wish to change it, after doing so, you need to also change the name
of the file in the line 11 of the code ('Students_grades.xlsx') to the name of your file
"""
import openpyxl as xl
import smtplib

doc = xl.load_workbook('Students_grades.xlsx')
# first sheet on the Exel document
sheet = doc['Sheet1']

# My email address and Your name represent the name and the email of the sender. In our case it is a professor
my_email_address = sheet.cell(2, 8).value
your_name = sheet.cell(2, 7).value
password = input('Enter your password ')

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(my_email_address, password)

# Accessing the different student's name and email address  present in the Excel document.
for row in range(2, sheet.max_row + 1):
    if sheet.cell(row, 2).value is None:
        continue
    student_name = sheet.cell(row, 1).value
    student_email = sheet.cell(row, 2).value
    grade = sheet.cell(row, 4).value
    letter_grade = sheet.cell(row, 5).value
    notes = sheet.cell(row, 3).value
    message = f'''
Good morning, {student_name}


I hope you are doing well. I just wanted to let you know that your grade has been released.
you got a {grade}/100 which is equivalent to an {letter_grade}.
if you have any question about your grades or anything, let me know. Thank you and Have a good one!


Best regards,
{your_name}

                '''
    if notes is not None:
        message = f'''
        Good morning, {student_name}


        I hope you are doing well. I just wanted to let you know that your grade has been released.
        you got a {grade}/100 which is equivalent to an {letter_grade}.
        if you have any question about your grades or anything, let me know. Thank you and Have a good one!
        NB: {notes}


        Best regards,
        {your_name}

                        '''
    server.sendmail(my_email_address, student_email, message)
    print('Message sent')
